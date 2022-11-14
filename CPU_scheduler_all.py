# reading a spreadsheet with cpu sceduling information appending the header and scedule into a seperate
# list later combining both into a dictionary
def read_spreadsheet(process):
    from openpyxl import load_workbook
    Workbook = load_workbook(filename="cpu-scheduling.xlsx")
    #Workbook = load_workbook(filename = "C:/Users/krist/OneDrive - Noroff Education AS/2_Operating Systems/Assesment 2/cpu-scheduling.xlsx")
    sheet = Workbook.active
    process = []
    header = ()
    for row in sheet.iter_rows(min_row = 1, max_row = 1, values_only = True):
        header = row

    for row in sheet.iter_rows(min_row = 2, values_only=True):
        temp = {header[0]:row[0], header[1]:row[1], header[2]:row[2], header[3]:row[3]}
        process.append(temp)
    return process


# Process based on Arrival time when process first arrives to cpu
def arrival_queue(process, process_queue, time_unit, algorithm):
    temp_queue = []
    temp_process_queue = process_queue
    for arrival_dict in process:
        if time_unit == arrival_dict['Arrival Time']:
            arrival_dict['Wait'] = 0
            temp_queue.append(arrival_dict)
    temp_queue = sort_queue(temp_queue, algorithm)
    for x in temp_queue:
        temp_process_queue.append(x)
    return temp_process_queue

# sort the queue depending on inputed sort strategy
def sort_queue(temp, sort_algorythm):
    priority_list = {"FIFO" : sorted(temp, key=lambda item: (item['Arrival Time'], item['Process ID'])),
     "SJF": sorted(temp, key=lambda item: (item['Instruction Load'],item['Arrival Time'], item['Process ID'])), 
     "PRI": sorted(temp, key=lambda item: (item['Priority'],item['Arrival Time'], item['Process ID'])), 
     "RR": sorted(temp, key=lambda item: (item['Arrival Time'], item['Instruction Load'], item['Process ID']))}
    if sort_algorythm == 'FIFO':
        temp1 = priority_list[sort_algorythm]
        return temp1
    
    elif sort_algorythm == 'SJF':
        temp1 = priority_list[sort_algorythm]
        return temp1

    elif sort_algorythm == 'PRI':    
        temp1 = priority_list[sort_algorythm]
        return temp1

    elif sort_algorythm == "RR":
        temp1 = priority_list[sort_algorythm]
        return temp1
    else:
        print("wrong input")

# responsible for printing in the specific structure when a proceses is being handeld
def print_queue(process_queue, time_unit, quantum, algorithm):
    if algorithm == "RR":
        temp_var = 0
        temp_var_1 = f"Q={quantum-1}"
    else:
        temp_var = -200
        temp_var_1 = ""

    if process_queue[0]['Instruction Load'] > 0 and quantum > temp_var:
        print(f"Time Unit {time_unit}: PID {process_queue[0]['Process ID']} executes. {process_queue[0]['Instruction Load']-1} instructions left. {temp_var_1}")
    elif process_queue[0]['Instruction Load'] == 0 or quantum == 0:
        print(f"Time Unit {time_unit}: Context switch.")      
  
# reposonible for print wait or no queue
def print_wait(process_queue):
    temp_print = []
    for print_item in process_queue[0:]:
        if len(print_item) > 0:
            temp_print.append(f"PID: {print_item['Process ID']} wait={print_item['Wait']}.")
    if len(process_queue[0:]) == 0:
        temp_print.append("No queue.")
    print(" ".join(temp_print))  
    print("\n")

# add process into the cpu
def update_process_list(cpu):
    cpu[0]['Instruction Load'] = cpu[0]['Instruction Load'] - 1
    return cpu

# updates the waittime of the process in the queue 
def update_wait_list(process_queue):
     for update_dict in process_queue[0:]:
        update_dict['Wait'] += 1
   
def idle_state(process):
    temp = 0
    for idle_state in process:
        if int(temp) < int(idle_state['Arrival Time']):
            temp = idle_state['Arrival Time']
    return temp

# handels the processes currently in the cpu, remove or appending depending on current state
def cpu_processing(temp, cpu, process_queue, stats):
    temp_stats = stats
    temp_cpu = cpu
    temp_process_queue = process_queue
    if temp == 1 and len(cpu) == 0:
        temp_cpu.append(process_queue.pop(0))
    elif temp == 2:
        temp_process_queue.append(temp_cpu.pop(0)) 
    elif temp == 3:
        temp_stats.append(temp_cpu.pop(0))
    return temp_cpu, temp_process_queue, temp_stats

# starts counting quantum when strategy chosen is RR
def algorithm_switch(algorithm, quantum):
    if algorithm == "RR":
        return 0
    else:
        quantum == 4
        return quantum


def main(temp):
    algorithm = temp
    process = []
    process = read_spreadsheet(process)    
    process_queue = []
    cpu = []
    quantum = 4
    time_unit = 0
    stats = []
    highest_arrvial_time = idle_state(process)
    if algorithm == "RR":
        this = 0
    else:
        this = -200

    while True:
        time_unit +=1
        process_queue = arrival_queue(process, process_queue, time_unit, algorithm)
        if algorithm == "SJF" or algorithm == "PRI":
            process_queue = sort_queue(process_queue, algorithm)

        if len(process_queue) > 0 or len(cpu) > 0:
            cpu_processing(1, cpu, process_queue, stats)
            if cpu[0]['Instruction Load'] > 0 and quantum > this:
                print_queue(cpu, time_unit, quantum,algorithm)
                update_process_list(cpu)
                cpu_processing(1, cpu, process_queue, stats)
                quantum -= 1
                update_wait_list(process_queue)
                print_wait(process_queue)
                        
            elif cpu[0]['Instruction Load'] == 0:
                print_queue(cpu, time_unit, quantum, algorithm)
                update_process_list(cpu)
                cpu_processing(3, cpu, process_queue, stats)
                quantum = 4
                update_wait_list(process_queue)
                print_wait(process_queue)
                
            elif algorithm == "RR" and cpu[0]['Instruction Load'] > 0 and quantum == 0:
                print_queue(cpu, time_unit, quantum, algorithm)
                cpu_processing(2, cpu, process_queue, stats)
                cpu_processing(1, cpu, process_queue, stats)
                quantum = 4
                update_wait_list(process_queue)
                print_wait(process_queue) 

        elif len(process_queue) == 0 and highest_arrvial_time > time_unit:
            print(f'Time Unit {time_unit}: CPU is idle.')
            print("\n")

        elif len(process_queue) == 0 and time_unit > highest_arrvial_time:
            print("All processes have executed. End of simulation")
            break

        else:
            print("OBS! Something went wrong that we didn't account for")



print('Type your sorting algorithm: "FIFO", "SJF", "PRI", "RR"')
temp = str(input()).upper()
print(f"You choose: {temp}")
print("\n")
main(temp)



